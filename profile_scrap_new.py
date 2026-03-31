import argparse
import json
import re
import time
import tkinter as tk
from pathlib import Path
from tkinter import simpledialog
from typing import Callable, Optional
from urllib.parse import parse_qs, quote_plus, unquote, urlparse
from xml.sax.saxutils import escape, quoteattr

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, TimeoutException, WebDriverException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from urllib3.exceptions import HTTPError, MaxRetryError, ProtocolError


DEFAULT_SEARCH_QUERY = "Mtech cse 27 iit Guwahati"
DEFAULT_MAX_PROFILES = 25
DEFAULT_OUTPUT_BASE = "results"
DEFAULT_HEADLESS = False
DEFAULT_CHALLENGE_TIMEOUT = 600

CHALLENGE_URL_FRAGMENTS = (
    "captcha",
    "/checkpoint",
    "/challenge",
    "challengeid",
    "/sorry/",
    "sorry/index",
)
CHALLENGE_TEXT_MARKERS = (
    "verify you are human",
    "security verification",
    "complete the security check",
    "unusual traffic from your computer network",
    "our systems have detected unusual traffic",
    "not a robot",
    "prove you're not a robot",
    "recaptcha",
    "hcaptcha",
)
CHALLENGE_IFRAME_SELECTORS = (
    "iframe[src*='recaptcha']",
    "iframe[src*='hcaptcha']",
    "iframe[src*='captcha']",
    "iframe[title*='CAPTCHA']",
    "iframe[title*='captcha']",
)
NAVIGATION_EXCEPTIONS = (
    ConnectionError,
    OSError,
    TimeoutException,
    WebDriverException,
    HTTPError,
    MaxRetryError,
    ProtocolError,
)


def normalize_whitespace(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def sanitize_filename_fragment(value: str) -> str:
    sanitized = re.sub(r"[^A-Za-z0-9._-]+", "_", value.strip())
    return sanitized.strip("._") or "results"


def prompt_runtime_settings(default_query: str, default_max_profiles: int) -> tuple[str, int]:
    try:
        root = tk.Tk()
    except tk.TclError:
        print("GUI prompt could not be opened. Falling back to the built-in query and max profile defaults.")
        return default_query, default_max_profiles

    root.withdraw()
    try:
        root.attributes("-topmost", True)
    except tk.TclError:
        pass

    try:
        query = simpledialog.askstring(
            "LinkedIn Link Scraper",
            "Enter the Google search query for LinkedIn profiles:",
            initialvalue=default_query,
            parent=root,
        )
        if query is None:
            query = default_query
        query = query.strip() or default_query

        max_profiles = simpledialog.askinteger(
            "LinkedIn Link Scraper",
            "Enter the maximum number of links to collect:",
            initialvalue=default_max_profiles,
            minvalue=1,
            parent=root,
        )
        if max_profiles is None:
            max_profiles = default_max_profiles

        return query, max_profiles
    finally:
        root.destroy()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Search Google for LinkedIn profile links and export them to PDF, JSON, and Excel."
    )
    parser.add_argument(
        "--query",
        default=DEFAULT_SEARCH_QUERY,
        help="Google query used with site:linkedin.com/in.",
    )
    parser.add_argument(
        "--max-profiles",
        type=int,
        default=DEFAULT_MAX_PROFILES,
        help="Maximum number of LinkedIn profile links to collect.",
    )
    parser.add_argument(
        "--output-base",
        default=DEFAULT_OUTPUT_BASE,
        help="Base output path without extension. Example: results/output_links",
    )
    parser.add_argument(
        "--headless",
        action=argparse.BooleanOptionalAction,
        default=DEFAULT_HEADLESS,
        help="Run Chrome in headless mode.",
    )
    parser.add_argument(
        "--challenge-timeout",
        type=int,
        default=DEFAULT_CHALLENGE_TIMEOUT,
        help="Seconds to wait for you to manually solve a CAPTCHA or checkpoint in the browser.",
    )
    args = parser.parse_args()
    args.query, args.max_profiles = prompt_runtime_settings(args.query, args.max_profiles)
    return args


def validate_args(args: argparse.Namespace) -> None:
    if not args.query:
        raise ValueError("The search query cannot be empty.")
    if args.max_profiles <= 0:
        raise ValueError("--max-profiles must be greater than 0.")
    if args.challenge_timeout <= 0:
        raise ValueError("--challenge-timeout must be greater than 0.")


def safe_current_url(driver: webdriver.Chrome) -> str:
    try:
        return driver.current_url or ""
    except WebDriverException:
        return ""


def safe_page_text(driver: webdriver.Chrome) -> str:
    try:
        body = driver.find_element(By.TAG_NAME, "body")
        return normalize_whitespace(body.text).lower()
    except (NoSuchElementException, WebDriverException):
        return ""


def detect_challenge_reason(driver: webdriver.Chrome) -> Optional[str]:
    current_url = safe_current_url(driver).lower()
    if any(fragment in current_url for fragment in CHALLENGE_URL_FRAGMENTS):
        return f"verification page at {current_url}"

    for selector in CHALLENGE_IFRAME_SELECTORS:
        try:
            if driver.find_elements(By.CSS_SELECTOR, selector):
                return "CAPTCHA iframe detected on the page"
        except WebDriverException:
            continue

    page_text = safe_page_text(driver)
    for marker in CHALLENGE_TEXT_MARKERS:
        if marker in page_text:
            return f"page text indicates verification: '{marker}'"

    return None


def wait_for_manual_challenge_resolution(
    driver: webdriver.Chrome,
    context: str,
    timeout_seconds: int,
    success_condition: Callable[[webdriver.Chrome], bool],
) -> bool:
    print(f"{context} requires manual verification in the Chrome window.")
    print(f"Solve it in the browser. Waiting up to {timeout_seconds} seconds before timing out.")

    deadline = time.time() + timeout_seconds
    last_reason: Optional[str] = None

    while time.time() < deadline:
        if success_condition(driver):
            print(f"{context} verification cleared. Resuming.")
            return True

        reason = detect_challenge_reason(driver)
        if reason and reason != last_reason:
            print(f"  Waiting on manual verification: {reason}")
            last_reason = reason
        elif not reason and last_reason:
            print("  Verification page looks cleared. Checking whether the run can continue...")
            last_reason = None

        time.sleep(2)

    print(f"{context} timed out after {timeout_seconds} seconds while waiting for manual verification.")
    return False


def ensure_verification_cleared(
    driver: webdriver.Chrome,
    context: str,
    timeout_seconds: int,
    headless: bool,
) -> bool:
    challenge_reason = detect_challenge_reason(driver)
    if not challenge_reason:
        return True

    print(f"{context} triggered a verification step: {challenge_reason}")
    if headless:
        print(f"{context} cannot be resolved manually while running headless. Rerun with --no-headless.")
        return False

    return wait_for_manual_challenge_resolution(
        driver,
        context,
        timeout_seconds,
        lambda current_driver: detect_challenge_reason(current_driver) is None,
    )


def wait_for_document(driver: webdriver.Chrome, timeout: int = 15) -> None:
    WebDriverWait(driver, timeout).until(
        lambda current_driver: current_driver.execute_script("return document.readyState") == "complete"
    )


def setup_driver(headless: bool = False) -> webdriver.Chrome:
    options = Options()
    options.add_argument("--start-maximized")
    options.add_argument("--disable-notifications")
    options.add_argument("--window-size=1600,1200")
    if headless:
        options.add_argument("--headless=new")

    driver = webdriver.Chrome(options=options)
    driver.set_page_load_timeout(60)
    return driver


def close_driver(driver: Optional[webdriver.Chrome]) -> None:
    if driver is None:
        return

    try:
        driver.quit()
    except Exception:
        pass


def restart_driver(driver: Optional[webdriver.Chrome], headless: bool, reason: str) -> webdriver.Chrome:
    print(f"Restarting Chrome because {reason}.")
    close_driver(driver)
    return setup_driver(headless=headless)


def navigate_with_retries(
    driver: webdriver.Chrome,
    url: str,
    context: str,
    headless: bool,
    max_attempts: int = 3,
) -> webdriver.Chrome:
    current_driver = driver
    last_exception: Optional[BaseException] = None

    for attempt in range(1, max_attempts + 1):
        try:
            if attempt > 1:
                print(f"{context}: retrying navigation ({attempt}/{max_attempts})")
            current_driver.get(url)
            wait_for_document(current_driver)
            return current_driver
        except KeyboardInterrupt:
            raise
        except NAVIGATION_EXCEPTIONS as exc:
            last_exception = exc
            print(f"{context}: navigation failed with {exc.__class__.__name__}: {exc}")

            if attempt >= max_attempts:
                break

            time.sleep(min(2 * attempt, 5))
            current_driver = restart_driver(
                current_driver,
                headless,
                f"{context.lower()} failed on attempt {attempt}",
            )

    raise RuntimeError(f"{context} failed after {max_attempts} navigation attempts.") from last_exception


def normalize_linkedin_profile_url(href: str) -> Optional[str]:
    if not href:
        return None

    parsed = urlparse(href)
    if "google." in parsed.netloc and parsed.path == "/url":
        href = parse_qs(parsed.query).get("q", [href])[0]

    href = unquote(href).strip()
    if not href:
        return None

    parsed = urlparse(href)
    if not parsed.scheme:
        parsed = urlparse(f"https://{href.lstrip('/')}")

    hostname = parsed.netloc.lower()
    if hostname.startswith("www."):
        hostname = hostname[4:]

    if hostname != "linkedin.com" and not hostname.endswith(".linkedin.com"):
        return None

    if not parsed.path.startswith("/in/"):
        return None

    clean_path = parsed.path.rstrip("/")
    return f"https://www.linkedin.com{clean_path}"


def find_next_google_page_button(driver: webdriver.Chrome):
    selectors = (
        (By.CSS_SELECTOR, "a#pnnext"),
        (By.CSS_SELECTOR, "a[aria-label='Next page']"),
        (By.XPATH, "//a[contains(., 'Next')]"),
    )
    for by, selector in selectors:
        elements = driver.find_elements(by, selector)
        for element in elements:
            if element.is_displayed():
                return element
    return None


def search_google_for_linkedin_profile_links(
    driver: webdriver.Chrome,
    search_query: str,
    max_results: int,
    headless: bool,
    challenge_timeout: int,
) -> tuple[webdriver.Chrome, list[str]]:
    linkedin_urls: list[str] = []
    google_query = f"site:linkedin.com/in {search_query}"
    google_search_url = f"https://www.google.com/search?q={quote_plus(google_query)}"

    print(f"Searching Google for LinkedIn profile links: {google_query}")
    driver = navigate_with_retries(driver, google_search_url, "Google search", headless)
    if not ensure_verification_cleared(driver, "Google search", challenge_timeout, headless):
        return driver, linkedin_urls

    page_num = 1
    while len(linkedin_urls) < max_results:
        if not ensure_verification_cleared(driver, f"Google search page {page_num}", challenge_timeout, headless):
            break

        print(f"\nScanning Google page {page_num}")
        WebDriverWait(driver, 15).until(lambda current_driver: bool(current_driver.find_elements(By.TAG_NAME, "a")))

        for link in driver.find_elements(By.TAG_NAME, "a"):
            try:
                normalized_url = normalize_linkedin_profile_url(link.get_attribute("href"))
            except WebDriverException:
                continue

            if normalized_url and normalized_url not in linkedin_urls:
                linkedin_urls.append(normalized_url)
                print(f"  + {normalized_url}")
                if len(linkedin_urls) >= max_results:
                    break

        if len(linkedin_urls) >= max_results:
            break

        next_button = find_next_google_page_button(driver)
        if next_button is None:
            print("No more Google result pages were found.")
            break

        current_url = safe_current_url(driver)
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", next_button)
        driver.execute_script("arguments[0].click();", next_button)
        WebDriverWait(driver, 15).until(lambda current_driver: safe_current_url(current_driver) != current_url)
        wait_for_document(driver)

        if not ensure_verification_cleared(driver, f"Google next page {page_num + 1}", challenge_timeout, headless):
            break

        page_num += 1

    print(f"\nFound {len(linkedin_urls)} unique LinkedIn profile links.")
    return driver, linkedin_urls


def resolve_output_base(output_base: str, search_query: str) -> Path:
    base_value = output_base.strip() if output_base else ""
    if not base_value:
        base_value = f"results_{sanitize_filename_fragment(search_query)}"

    base_path = Path(base_value).expanduser()
    if base_path.suffix:
        base_path = base_path.with_suffix("")

    if not base_path.is_absolute():
        base_path = Path.cwd() / base_path

    base_path.parent.mkdir(parents=True, exist_ok=True)
    return base_path


def save_links_to_json(links: list[str], search_query: str, output_base: Path) -> Path:
    output_path = output_base.with_suffix(".json")
    payload = {
        "search_query": search_query,
        "total_links": len(links),
        "links": links,
    }
    output_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    return output_path


def save_links_to_excel(links: list[str], search_query: str, output_base: Path) -> Path:
    output_path = output_base.with_suffix(".xlsx")
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "LinkedIn Links"

    headers = ["Search Query", "Index", "LinkedIn URL"]
    for column, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=column)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_index, url in enumerate(links, start=2):
        worksheet.cell(row=row_index, column=1, value=search_query)
        worksheet.cell(row=row_index, column=2, value=row_index - 1)
        url_cell = worksheet.cell(row=row_index, column=3, value=url)
        url_cell.hyperlink = url
        url_cell.style = "Hyperlink"
        url_cell.alignment = Alignment(horizontal="left", vertical="center")

    worksheet.column_dimensions["A"].width = 40
    worksheet.column_dimensions["B"].width = 10
    worksheet.column_dimensions["C"].width = 90

    workbook.save(output_path)
    return output_path


def save_links_to_pdf(links: list[str], search_query: str, output_base: Path) -> Path:
    output_path = output_base.with_suffix(".pdf")

    document = SimpleDocTemplate(
        str(output_path),
        pagesize=A4,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36,
    )
    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    meta_style = styles["BodyText"]
    link_style = ParagraphStyle(
        "LinkStyle",
        parent=styles["BodyText"],
        fontSize=10,
        leading=14,
        textColor=colors.HexColor("#0B63CE"),
        spaceAfter=8,
    )

    story = [
        Paragraph("LinkedIn Profile Links", title_style),
        Spacer(1, 0.15 * inch),
        Paragraph(f"Search query: {escape(search_query)}", meta_style),
        Paragraph(f"Total links: {len(links)}", meta_style),
        Spacer(1, 0.2 * inch),
    ]

    for index, url in enumerate(links, start=1):
        href = quoteattr(url)
        story.append(Paragraph(f"{index}. <link href={href}>{escape(url)}</link>", link_style))

    document.build(story)
    return output_path


def main() -> None:
    args = parse_args()
    validate_args(args)

    driver: Optional[webdriver.Chrome] = None
    try:
        driver = setup_driver(headless=args.headless)

        print("\n" + "=" * 70)
        print("STEP 1: COLLECTING LINKEDIN PROFILE LINKS")
        print("=" * 70)
        driver, linkedin_links = search_google_for_linkedin_profile_links(
            driver,
            args.query,
            args.max_profiles,
            args.headless,
            args.challenge_timeout,
        )

        if not linkedin_links:
            print("No LinkedIn links were found. Try a different query.")
            return

        output_base = resolve_output_base(args.output_base, args.query)
        json_path = save_links_to_json(linkedin_links, args.query, output_base)
        excel_path = save_links_to_excel(linkedin_links, args.query, output_base)
        pdf_path = save_links_to_pdf(linkedin_links, args.query, output_base)

        print("\n" + "=" * 70)
        print("EXPORT COMPLETE")
        print("=" * 70)
        print(f"Total links exported: {len(linkedin_links)}")
        print(f"JSON: {json_path}")
        print(f"Excel: {excel_path}")
        print(f"PDF: {pdf_path}")

    except Exception as exc:
        print(f"Error in main execution: {exc}")
        raise
    finally:
        close_driver(driver)
        print("\nBrowser closed")


if __name__ == "__main__":
    main()

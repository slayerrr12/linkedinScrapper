import json
from pathlib import Path
from xml.sax.saxutils import escape, quoteattr

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer

from .utils import sanitize_filename_fragment


def resolve_output_base(output_base: str, search_query: str) -> Path:
    base_value = output_base.strip() if output_base else ""
    if not base_value:
        base_value = f"results_{sanitize_filename_fragment(search_query)}"

    base_path = Path(base_value).expanduser()
    if base_path.suffix:
        base_path = base_path.with_suffix("")

    if not base_path.is_absolute():
        base_path = Path.cwd() / base_path

    base_path.parent.mkdir(parents=True, exist_ok=True)
    return base_path


def save_links_to_json(links: list[str], search_query: str, output_base: Path) -> Path:
    output_path = output_base.with_suffix(".json")
    payload = {
        "search_query": search_query,
        "total_links": len(links),
        "links": links,
    }
    output_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    return output_path


def save_links_to_excel(links: list[str], search_query: str, output_base: Path) -> Path:
    output_path = output_base.with_suffix(".xlsx")
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "LinkedIn Links"

    headers = ["Search Query", "Index", "LinkedIn URL"]
    for column, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=column)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_index, url in enumerate(links, start=2):
        worksheet.cell(row=row_index, column=1, value=search_query)
        worksheet.cell(row=row_index, column=2, value=row_index - 1)
        url_cell = worksheet.cell(row=row_index, column=3, value=url)
        url_cell.hyperlink = url
        url_cell.style = "Hyperlink"
        url_cell.alignment = Alignment(horizontal="left", vertical="center")

    worksheet.column_dimensions["A"].width = 40
    worksheet.column_dimensions["B"].width = 10
    worksheet.column_dimensions["C"].width = 90

    workbook.save(output_path)
    return output_path


def save_links_to_pdf(links: list[str], search_query: str, output_base: Path) -> Path:
    output_path = output_base.with_suffix(".pdf")

    document = SimpleDocTemplate(
        str(output_path),
        pagesize=A4,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36,
    )
    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    meta_style = styles["BodyText"]
    link_style = ParagraphStyle(
        "LinkStyle",
        parent=styles["BodyText"],
        fontSize=10,
        leading=14,
        textColor=colors.HexColor("#0B63CE"),
        spaceAfter=8,
    )

    story = [
        Paragraph("LinkedIn Profile Links", title_style),
        Spacer(1, 0.15 * inch),
        Paragraph(f"Search query: {escape(search_query)}", meta_style),
        Paragraph(f"Total links: {len(links)}", meta_style),
        Spacer(1, 0.2 * inch),
    ]

    for index, url in enumerate(links, start=1):
        href = quoteattr(url)
        story.append(Paragraph(f"{index}. <link href={href}>{escape(url)}</link>", link_style))

    document.build(story)
    return output_path


def export_links(links: list[str], search_query: str, output_base: str) -> dict[str, Path]:
    base_path = resolve_output_base(output_base, search_query)
    return {
        "json": save_links_to_json(links, search_query, base_path),
        "excel": save_links_to_excel(links, search_query, base_path),
        "pdf": save_links_to_pdf(links, search_query, base_path),
    }
